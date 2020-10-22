from openpyxl import load_workbook


def get_all_row(path):
    workbook = load_workbook(path)
    booksheet = workbook.active
    return booksheet.max_row

def get_all_lolumn(path):
    workbook = load_workbook(path)
    booksheet = workbook.active
    return booksheet.max_column

def get_row(h,path):
    workbook = load_workbook(path)
    booksheet = workbook.active
    lp=3
    result=[]
    for row in range(1,lp+1):
        cell_data_1 = booksheet.cell(row=h, column=row).value
        result.append(cell_data_1)
    return result

def read(path):
    list=[]
    for i in range(2,get_all_row(path)+1):
        a = get_row(i, path)
        d={
            "name":a[0],
            "age":a[1],
            "address":a[2],
        }
        list.append(d)
    return list

def get_write(data):
    import pymongo
    print(data)
    client=pymongo.MongoClient('localhost',27017)
    ceshi=client["Mydb"]
    zhaoping=ceshi["mydata"]
    for i in data:
        print(i)
        zhaoping.insert_one(i)
    return True